"""Tests for Excel and HTML export."""


import openpyxl

from staff_duty.export import write_excel, write_html


class TestExcelExport:
    """Verify Excel workbook structure."""

    def test_excel_has_expected_sheets(self, solved_roster, tmp_path):
        """Roster, Summary, Fairness sheets exist."""
        sdnco_sol, runner_sol, all_days, holiday_dates = solved_roster
        out = tmp_path / "test.xlsx"
        write_excel([sdnco_sol, runner_sol], all_days, holiday_dates, out)

        wb = openpyxl.load_workbook(out)
        sheet_names = wb.sheetnames
        assert "Roster" in sheet_names
        assert "Summary" in sheet_names
        assert "Fairness" in sheet_names
        wb.close()

    def test_excel_roster_row_count(self, solved_roster, tmp_path):
        """Roster sheet row count matches number of days (plus header)."""
        sdnco_sol, runner_sol, all_days, holiday_dates = solved_roster
        out = tmp_path / "test.xlsx"
        write_excel([sdnco_sol, runner_sol], all_days, holiday_dates, out)

        wb = openpyxl.load_workbook(out)
        ws = wb["Roster"]
        # Data rows = total rows minus header row(s)
        data_rows = ws.max_row - 1  # at least 1 header row
        assert data_rows >= len(all_days), (
            f"Expected at least {len(all_days)} data rows, got {data_rows}"
        )
        wb.close()


class TestHtmlExport:
    """Verify HTML dashboard output."""

    def test_html_contains_table(self, solved_roster, tmp_path):
        """Output HTML has cal-table class."""
        sdnco_sol, runner_sol, all_days, holiday_dates = solved_roster
        out = tmp_path / "test.html"
        write_html([sdnco_sol, runner_sol], all_days, holiday_dates, out)

        html = out.read_text(encoding="utf-8")
        assert "cal-table" in html

    def test_html_contains_all_directorates(self, solved_roster, tmp_path):
        """All directorate names appear in the HTML output."""
        sdnco_sol, runner_sol, all_days, holiday_dates = solved_roster
        out = tmp_path / "test.html"
        write_html([sdnco_sol, runner_sol], all_days, holiday_dates, out)

        html = out.read_text(encoding="utf-8")
        for dir_name in ["G1", "G2", "G3"]:
            assert dir_name in html, f"Directorate {dir_name} not found in HTML"
