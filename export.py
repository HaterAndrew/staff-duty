"""
Export roster solutions to Excel (.xlsx) and static HTML dashboard.

Excel workbook structure (one file, two roles side-by-side):
  Sheet "Roster"   — calendar table with one row per day; SDNCO and SD_Runner columns
  Sheet "Summary"  — per-directorate stats for each role
  Sheet "Fairness" — Gini coefficients + target vs. actual comparison

HTML dashboard:
  - USAREUR-AF HHBn branding (navy + gold, inline USAREUR insignia SVG)
  - Tabular list calendar — one row per day, SDNCO and SD_Runner columns side-by-side
  - Summary table per role with inline bar chart
  - Fairness analysis with Gini + target vs. actual delta
"""

from __future__ import annotations

import math
from datetime import date
from pathlib import Path
from typing import Dict, List, Set

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

from .calendar_utils import HOLIDAY, WEEKEND, WEEKDAY, classify_day
from .solver import DirectorateStats, RosterSolution


# ── Colour palette (max 16 directorates) ─────────────────────────────────────
_PALETTE = [
    "4472C4", "ED7D31", "A9D18E", "FF0000", "FFC000",
    "00B0F0", "7030A0", "92D050", "FF00FF", "00B050",
    "C00000", "BDD7EE", "F4B183", "D9E1F2", "FFE699",
    "C6E0B4",
]


def _dir_color_map(all_dirs: List[str]) -> Dict[str, str]:
    return {dn: _PALETTE[i % len(_PALETTE)] for i, dn in enumerate(all_dirs)}


# ── Excel export ─────────────────────────────────────────────────────────────

def write_excel(
    solutions: List[RosterSolution],
    all_days: List[date],
    holiday_dates: Set[date],
    output_path: Path,
) -> None:
    """
    Write a single Excel workbook containing all provided solutions.
    Typically two solutions: SDNCO and SD_Runner.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # delete default sheet

    all_dirs = _all_dir_names(solutions)
    color_map = _dir_color_map(all_dirs)

    _write_roster_sheet(wb, solutions, all_days, holiday_dates, color_map)
    _write_summary_sheet(wb, solutions)
    _write_fairness_sheet(wb, solutions)

    wb.save(output_path)


def _all_dir_names(solutions: List[RosterSolution]) -> List[str]:
    seen, names = set(), []
    for sol in solutions:
        for s in sol.stats:
            if s.name not in seen:
                seen.add(s.name)
                names.append(s.name)
    return names


def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _write_roster_sheet(
    wb: openpyxl.Workbook,
    solutions: List[RosterSolution],
    all_days: List[date],
    holiday_dates: Set[date],
    color_map: Dict[str, str],
) -> None:
    """
    Working roster sheet — designed to be printed and filled in.

    Layout (one row per day):
      Date | Day | Type | [ROLE — UNIT | ROLE — NAME (blank)] × n_solutions

    Name columns are light-yellow with a dotted border — ready for handwriting
    or typing of individual soldier names. Month-start rows are bold to visually
    separate months. Print settings: landscape, repeat header rows, fit to width.
    """
    ws = wb.create_sheet("Roster")

    n_roles  = len(solutions)
    n_cols   = 3 + n_roles * 2     # 3 fixed + (Unit + Name) per role

    # ── Row 1 — document title ─────────────────────────────────────────────────
    q_start = all_days[0]
    q_end   = all_days[-1]
    title   = (
        f"ASCC HQ STAFF DUTY ROSTER  \u2014  "
        f"{q_start.strftime('%d %b %Y')} \u2013 {q_end.strftime('%d %b %Y')}"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font      = Font(bold=True, size=13, color="FFFFFF")
    tc.fill      = _header_fill("0C2340")   # USAREUR-AF navy
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2 — column headers ─────────────────────────────────────────────────
    fixed_hdrs = ["DATE", "DAY", "TYPE"]
    role_hdrs  = []
    for sol in solutions:
        role_hdrs += [f"{sol.role}\nUNIT", f"{sol.role}\nNAME"]
    headers = fixed_hdrs + role_hdrs

    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = _header_fill("1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()
    ws.row_dimensions[2].height = 30

    # ── Fill definitions ───────────────────────────────────────────────────────
    _WEEKEND_FILL = PatternFill("solid", fgColor="DDEBF7")
    _HOLIDAY_FILL = PatternFill("solid", fgColor="FCE4D6")
    _WEEKDAY_FILL = PatternFill("solid", fgColor="FFFFFF")

    # Name-column fills — light yellow signals "fill me in"
    _NAME_FILL = {
        WEEKDAY: PatternFill("solid", fgColor="FFFACD"),
        WEEKEND: PatternFill("solid", fgColor="FFF3B0"),
        HOLIDAY: PatternFill("solid", fgColor="FFEEA0"),
    }
    _NAME_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="dotted"), bottom=Side(style="dotted"),
    )

    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    # ── Data rows (start at row 3) ─────────────────────────────────────────────
    for ri_offset, day in enumerate(all_days):
        ri       = ri_offset + 3
        day_type = classify_day(day, holiday_dates)
        row_fill = {HOLIDAY: _HOLIDAY_FILL, WEEKEND: _WEEKEND_FILL, WEEKDAY: _WEEKDAY_FILL}[day_type]

        # Date — bold on 1st of month (visual month separator)
        date_cell = ws.cell(row=ri, column=1, value=day.strftime("%d %b %Y"))
        date_cell.fill      = row_fill
        date_cell.border    = _thin_border()
        date_cell.alignment = Alignment(horizontal="center")
        if day.day == 1:
            date_cell.font = Font(bold=True)

        day_cell = ws.cell(row=ri, column=2, value=day_names[day.weekday()])
        day_cell.fill      = row_fill
        day_cell.border    = _thin_border()
        day_cell.alignment = Alignment(horizontal="center")

        type_cell = ws.cell(row=ri, column=3, value=day_type.title())
        type_cell.fill      = row_fill
        type_cell.border    = _thin_border()
        type_cell.alignment = Alignment(horizontal="center")

        # Role columns: Unit (colored) + Name (blank, yellow)
        for sol_idx, sol in enumerate(solutions):
            unit_col = 4 + sol_idx * 2
            name_col = unit_col + 1

            assigned  = sol.assignment.get(day, "—")
            unit_cell = ws.cell(row=ri, column=unit_col, value=assigned)
            unit_cell.alignment = Alignment(horizontal="center")
            unit_cell.border    = _thin_border()
            if assigned in color_map:
                unit_cell.fill = PatternFill("solid", fgColor=color_map[assigned])
            else:
                unit_cell.fill = row_fill

            name_cell = ws.cell(row=ri, column=name_col, value="")
            name_cell.fill      = _NAME_FILL[day_type]
            name_cell.border    = _NAME_BORDER
            name_cell.alignment = Alignment(horizontal="left")

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 13   # Date
    ws.column_dimensions["B"].width = 5    # Day
    ws.column_dimensions["C"].width = 10   # Type
    for sol_idx in range(n_roles):
        unit_col = 4 + sol_idx * 2
        name_col = unit_col + 1
        ws.column_dimensions[get_column_letter(unit_col)].width = 15   # Unit
        ws.column_dimensions[get_column_letter(name_col)].width = 24   # Name — wide for text

    # ── Freeze header rows ─────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Print settings ─────────────────────────────────────────────────────────
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0       # unlimited pages tall
    ws.page_setup.fitToPage    = True
    ws.page_margins.left       = 0.5
    ws.page_margins.right      = 0.5
    ws.page_margins.top        = 0.75
    ws.page_margins.bottom     = 0.75
    ws.print_title_rows        = "1:2"   # repeat title + header on every page

    last_row = len(all_days) + 2
    ws.print_area = f"A1:{get_column_letter(n_cols)}{last_row}"

    ws.oddHeader.center.text = "&B&12ASCC HQ STAFF DUTY ROSTER"
    ws.oddFooter.left.text   = "UNCLASSIFIED"
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.right.text  = "Generated: &D"


def _write_summary_sheet(
    wb: openpyxl.Workbook,
    solutions: List[RosterSolution],
) -> None:
    ws = wb.create_sheet("Summary")
    row = 1

    for sol in solutions:
        # Role title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        title = ws.cell(row=row, column=1, value=f"Role: {sol.role}  |  Solver: {sol.solver_status}")
        title.font = Font(bold=True, size=12, color="FFFFFF")
        title.fill = _header_fill("2E75B6")
        title.alignment = Alignment(horizontal="left")
        row += 1

        headers = ["Directorate", "Eligible", "Total Days", "Weekday", "Weekend", "Holiday", "Hard Days"]
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True)
            c.fill = _header_fill("BDD7EE")
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()
        row += 1

        for s in sol.stats:
            vals = [s.name, s.eligible, s.total_days, s.weekday_days, s.weekend_days, s.holiday_days, s.hard_days]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=ci, value=v)
                c.alignment = Alignment(horizontal="center")
                c.border = _thin_border()
            row += 1

        row += 2  # blank separator

    for ci in range(1, 8):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _write_fairness_sheet(
    wb: openpyxl.Workbook,
    solutions: List[RosterSolution],
) -> None:
    ws = wb.create_sheet("Fairness")
    row = 1

    for sol in solutions:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        title = ws.cell(row=row, column=1, value=f"Fairness — {sol.role}")
        title.font = Font(bold=True, size=12, color="FFFFFF")
        title.fill = _header_fill("2E75B6")
        row += 1

        gini_row = ws.cell(row=row, column=1, value="Gini Coefficient (0=fair, 1=unfair)")
        gini_row.font = Font(italic=True)
        ws.cell(row=row, column=2, value=f"Total days: {sol.total_day_gini:.4f}")
        ws.cell(row=row, column=3, value=f"Hard days: {sol.hard_day_gini:.4f}")
        row += 2

        headers = ["Directorate", "Eligible", "Total Target", "Total Actual", "Hard Target", "Hard Actual"]
        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(bold=True)
            c.fill = _header_fill("BDD7EE")
            c.alignment = Alignment(horizontal="center")
            c.border = _thin_border()
        row += 1

        H = sum(s.eligible for s in sol.stats)
        n = len(sol.assignment)
        n_hard = sum(1 for s in sol.stats for _ in range(s.hard_days))  # total hard days

        # Recompute n_hard from stats
        n_hard = sum(s.hard_days for s in sol.stats)

        for s in sol.stats:
            total_target = s.eligible / H * n
            hard_target  = s.eligible / H * n_hard
            vals = [
                s.name, s.eligible,
                round(total_target, 1), s.total_days,
                round(hard_target, 1), s.hard_days,
            ]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=ci, value=v)
                c.alignment = Alignment(horizontal="center")
                c.border = _thin_border()

                # Colour the "Actual" cells based on deviation
                if ci in (4, 6):
                    target_col = {4: 3, 6: 5}[ci]
                    target_val = vals[target_col - 1]
                    delta = abs(v - target_val)
                    if delta <= 1:
                        c.fill = PatternFill("solid", fgColor="C6EFCE")  # green — within ±1
                    elif delta <= 2:
                        c.fill = PatternFill("solid", fgColor="FFEB9C")  # yellow
                    else:
                        c.fill = PatternFill("solid", fgColor="FFC7CE")  # red

            row += 1

        row += 2

    for ci in range(1, 7):
        ws.column_dimensions[get_column_letter(ci)].width = 16


# ── USAREUR-AF insignia SVG (inline — no external file dependency) ─────────────
_USAREUR_SVG = '''<svg width="60" height="79" viewBox="0 0 246.08 322.7" xmlns="http://www.w3.org/2000/svg">
<path d="m123.04 1.2002c-66.774-0.10337-121.84 36.129-121.84 36.129s-0.41936 102.81 31.635 167.86c38.542 78.221 87.955 114.53 90.014 116.02v0.28711s0.17836-0.13656 0.18945-0.14453c0.0111 8e-3 0.18945 0.14453 0.18945 0.14453v-0.28711c2.0583-1.491 51.471-37.799 90.014-116.02 32.054-65.053 31.635-167.86 31.635-167.86s-55.064-36.232-121.84-36.129z" fill="#171695"/>
<path d="m117.65 200.15c0.22276-24.163 0.72429-48.331 0.76474-72.491-3.4093-2.716-6.1771-7.05-5.7856-11.546 1.9789-5.9597 1.7443-13.803-3.9582-17.636 1.8433 5.0171 0.26645 10.325-2.0321 14.893-3.226 6.8124-3.2514 15.826 2.3417 21.48 3.7657 3.5442 4.8942 8.9284 4.1338 13.888-0.90701-4.1592-3.1854-8.269-7.0592-10.453-2.6858-1.4232-6.0935-2.2012-7.68-5.4303-1.8455-4.6228 0.57719-10.883-3.664-14.655-2.1153-2.7543-0.60992 2.8488-1.4523 6.005-2.7227 6.9975-2.736 16.688 3.9518 21.421 7.0203 3.051 11.93 10.347 14.244 17.644 0.64053 2.8573 3.6353 9.829-1.026 5.0949-2.0867-1.4452-4.7634-1.5364-7.2033-2.196-2.2111-0.97619-4.6484-2.1535-5.8647-4.6307-1.2525-6.7559-3.6439 3.0959-2.5354 5.7975 0.75012 3.2858 3.5415 5.8686 6.5368 6.7892 5.3657 2.4602 10.013 7.6601 11.416 13.658 1.4895 3.9304 0.0343 1.8767-2.3495 1.4882-4.1636-0.67852-8.7009-1.1125-8.6626-6.3935-1.5588 2.6439-0.35735 9.9462 3.2674 12.781 4.1804 3.9253 8.5014 8.6678 7.7912 14.848-1.3426-1.0678-3.4449-6.4157-6.6622-8.0086-2.5913-1.283-5.5194-1.6909-8.4267-1.92-4.2628-0.33596-5.5236-7.0758-6.278-8.16-2.1353 6.2842-1.7474 14.361 4.1447 18.4 1.8745 1.3902 3.8662 1.9662 5.76 3.2 6.1234 3.616 12.541 9.1598 12.866 16.876-2.1926-1.5176-4.359-1.9963-6.8931-2.5488-5.1244-1.1616-6.3828-3.2136-7.6173-5.6054-1.4068 3.7385-1.4076 9.087 2.604 11.332 2.3092 1.2364 5.0499 1.6519 7.36 2.7733 3.2903 1.5294 6.17 4.5116 7.5368 7.9542 0.16138-14.883 0.24569-29.766 0.43081-44.648zm11.787 44.36c2.0528-3.4325 5.9175-6.1306 9.2582-7.6665 3.1644-1.7088 5.4601-5.2564 6.4126-8.8633 0.9782-2.5496 3.1406-4.6863 5.5341-5.5192-1.4595-2.3635-7.7145-1.0161-8.9744 2.2226-1.2402 2.1941-2.0186 6.2395-4.8538 6.6002-1.8386-4.7818 1.8391-9.3228 5.7216-11.72 4.751-2.7173 5.9403-8.4822 5.6681-13.561-0.47323-2.6628 3.5956-7.4376-0.97477-4.7862-3.3806 2.1948-3.9949 6.2219-6.4596 8.96-1.7323 1.2846-6.6189 5.1661-4.9206 0.2601-0.15386-6.2165 6.2318-9.3808 7.6425-15.009 2.0182-5.3841 1.3992-12.149 6.3626-16.003 1.1581-1.3309 3.7292-1.4513 4.4717-2.4653-2.6409-0.81956-5.6513-1.9782-8.4833-0.916-4.0135 0.9504-5.8123 5.6086-6.8508 9.3521-0.84516 2.3388-3.826 8.1323-5.3123 6.7794-0.73094-5.4087 0.42272-11.13 4.3828-15.052 3.1556-3.1139 7.6264-5.4018 7.7325-10.478 0.85585-5.8424 0.24119-12.857 5.1679-17.054 1.2324-1.6278 4.8392-1.7365 5.3333-2.56-2.3741-1.4543-5.6625-2.2649-8.5333-1.28-4.7492 1.533-8.2315 6.7727-8.8639 11.774-0.15 2.4524-3.3149 9.591-4.6563 7.583 0.0931-7.8592 0.74192-16.675 6.5869-22.549 3.647-4.2573 4.6802-11.032 2.1333-16.008-3.2221-5.4172-2.8943-12.673-0.10357-18.353 0.33896-1.8011-3.7282 0.34225-4.8465 1.1494-4.6125 3.5553-2.6036 9.7857-2.6381 14.684-0.26023 4.4475-2.3158 9.4809-6.7049 11.115-1.8892 0.32158-0.70904 2.7968-1.0036 4.3335 0.47611 35.043 0.25317 70.091 0.97152 105.13 0.0771 2.5847 0.077 1.0616 0.80023-0.0997z" fill="#ba122b"/>
<path d="m122.87 85.927-0.17773 0.3125-3.8711 8.1992-1.6973 152.51c3.8015 0.6784 7.7755 0.7862 11.584 0.14844l-1.3906-152.44z" fill="#fff"/>
<path d="m89.126 237.58c-1.8717 0.0334-3.9322 0.58716-5.0742 1.4277-1.3664 1.0057-2.508 3.7851-1 4.5625 20.475 10.556 33.701 9.6191 33.701 9.6191l-0.95117 33.131s0.27688 3.1109 6.25 3.25c8.623 0.2008 8.125-3.875 8.125-3.875l-0.64844-32.506c2.296 0.0577 14.762 0.0804 33.221-9.4356 1.508-0.77743 0.36639-3.5568-1-4.5625-1.8272-1.3449-6.0058-1.9571-8-0.875-13.247 7.1882-15.999 7.6085-25.029 9.2227-4e-3-0.0259 9e-5-8.9e-4-4e-3-0.0274-3.8141 0.63948-7.7946 0.53021-11.602-0.15039-9.0588-1.6195-11.798-2.0308-25.062-9.2285-0.74781-0.40578-1.8028-0.57276-2.9258-0.55274z" fill="#ffc61d"/>
<path d="m124.16 13.198c-52.79-0.68772-108.32 30.463-109.89 31.35 1.0035 14.131 1.8484 28.233 3.043 42.281 6.1467-5.5066 14.782-12.539 25.912-19.244 19.397-11.685 46.408-22.446 80.895-22.699h0.043c34.486 0.25299 61.498 11.014 80.895 22.699 10.35 6.2354 18.505 12.731 24.551 18.045 1.0598-13.763 2.0064-28.18 2.6152-40.85-1.5242-0.87987-53.825-30.875-108.06-31.582z" fill="#8c99a2"/>
<path d="m124.12 44.886c-34.486 0.25299-61.498 11.014-80.895 22.699-11.132 6.7062-19.761 13.734-25.912 19.244 0.21312 2.5063 0.44094 5.0111 0.67969 7.5137 6.0248-5.6655 15.323-13.783 28.328-21.617 18.681-11.254 44.567-21.591 77.82-21.84 33.253 0.24842 59.139 10.585 77.82 21.84 12.097 7.2878 21.008 14.837 27.049 20.42 0.20304-2.4866 0.40142-4.9768 0.59766-7.5254-6.0497-5.3122-14.226-11.815-24.551-18.035-19.397-11.685-46.408-22.446-80.895-22.699h-0.0215z" fill="#ba122b"/>
<path d="m124.11 62.792v2e-3c-56.492 0.0247-91.747 32.586-104.46 46.791 0.32892 2.6537 0.67816 5.3047 1.0566 7.9531 7.5488-9.4708 42.508-48.73 103.44-48.744 59.368 0.01351 94.021 37.235 102.74 47.906 0.26319-2.6049 0.52829-5.3038 0.79492-8.1777-13.495-14.715-48.42-45.704-103.51-45.729v-2e-3z" fill="#144d2a"/>
<path d="m124.11 56.761v2e-3c-56.714 0.0244-93.28 32.374-105.36 44.945 0.28465 2.6373 0.57198 5.2738 0.89844 7.9062 4.9499-5.8673 41.751-46.838 104.49-46.852 59.67 0.01333 95.878 37.077 103.54 45.754 0.23826-2.5683 0.47617-5.2212 0.71289-7.9551-13.065-13.255-49.088-43.775-104.22-43.799v-2e-3z" fill="#fcc916"/>
<path d="m124.11 50.886v0.0039c-54.122-0.06546-90.709 28.355-106.13 43.201 0.24278 2.5552 0.51159 5.1072 0.78711 7.6582 11.989-12.516 48.534-44.948 105.38-44.861 55.264-0.0845 91.263 30.505 104.25 43.713 0.21785-2.5151 0.43124-5.1147 0.64453-7.7324-16.122-15.131-52.202-42.042-104.86-41.979v-0.0039z" fill="#fa9e0d"/>
</svg>'''


# ── HTML export ───────────────────────────────────────────────────────────────

def write_html(
    solutions: List[RosterSolution],
    all_days: List[date],
    holiday_dates: Set[date],
    output_path: Path,
) -> None:
    """Write a standalone HTML dashboard with USAREUR-AF HHBn branding."""
    all_dirs = _all_dir_names(solutions)
    color_map = _dir_color_map(all_dirs)

    quarter_label = f"{all_days[0].strftime('%d %b %Y')} – {all_days[-1].strftime('%d %b %Y')}"
    generated    = date.today().strftime("%d %b %Y")
    n_days       = len(all_days)
    n_hard       = sum(1 for d in all_days
                       if classify_day(d, holiday_dates) in (WEEKEND, HOLIDAY))
    n_weekday    = n_days - n_hard

    stat_cards = f"""
<div class="stat-row">
  <div class="stat-card"><div class="stat-val">{n_days}</div><div class="stat-lbl">Total Days</div></div>
  <div class="stat-card"><div class="stat-val">{n_weekday}</div><div class="stat-lbl">Weekdays</div></div>
  <div class="stat-card"><div class="stat-val">{n_hard}</div><div class="stat-lbl">Wknd / Holiday</div></div>
  <div class="stat-card"><div class="stat-val">{len(solutions[0].stats) if solutions else 0}</div><div class="stat-lbl">Directorates</div></div>
</div>"""

    legend_html   = _build_legend_html(all_dirs, color_map)
    calendar_html = _build_calendar_table_html(solutions, all_days, holiday_dates, color_map)
    summary_html  = _build_summary_html(solutions, all_days, color_map)
    fairness_html = _build_fairness_html(solutions)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>HHBn Staff Duty Roster — {quarter_label}</title>
<style>
/* ── Design tokens (USAREUR-AF Command colours) ──────────────────── */
:root {{
  --navy:        #0C2340;
  --navy-dark:   #071628;
  --navy-light:  #163A6C;
  --navy-mid:    #1E4A88;
  --navy-pale:   #EEF2FA;
  --gold:        #C8971A;
  --gold-light:  #E0B840;
  --gold-dark:   #9A7010;
  --gold-pale:   #FDF5DC;
  --white:       #FFFFFF;
  --off-white:   #F3F5FA;
  --gray-50:     #EFF1F8;
  --gray-100:    #E0E4EF;
  --gray-200:    #C4CAE0;
  --gray-400:    #7A88A8;
  --gray-600:    #485878;
  --gray-700:    #303C58;
  --gray-900:    #0A1628;
  --weekend-bg:  #DBEAFE;
  --holiday-bg:  #FCE4D6;
  --green-ok:    #166534;
  --green-bg:    #DCFCE7;
  --yellow-ok:   #854D0E;
  --yellow-bg:   #FEF9C3;
  --red-ok:      #991B1B;
  --red-bg:      #FEE2E2;
  --radius:      3px;
  --radius-lg:   6px;
  --shadow:      0 1px 5px rgba(0,0,0,0.10);
  --shadow-hdr:  0 3px 16px rgba(0,0,0,0.35);
  --font-body:   Inter, 'Segoe UI', system-ui, Arial, sans-serif;
  --font-ui:     Arial, Helvetica, sans-serif;
  --font-mono:   'Courier New', Courier, monospace;
}}

/* ── Reset ───────────────────────────────────────────────────────── */
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: var(--font-body);
  background: var(--off-white);
  color: var(--gray-900);
  font-size: 14px;
  line-height: 1.6;
  min-height: 100vh;
}}

/* ── Classification banner ───────────────────────────────────────── */
.cls-banner {{
  background: #1E6B2A;
  color: #fff;
  text-align: center;
  font-family: var(--font-ui);
  font-weight: 700;
  font-size: 11px;
  letter-spacing: 4px;
  padding: 5px 0;
  text-transform: uppercase;
}}

/* ── Header ──────────────────────────────────────────────────────── */
header {{
  background: linear-gradient(155deg, var(--navy-dark) 0%, var(--navy) 55%, var(--navy-light) 100%);
  color: var(--white);
  box-shadow: var(--shadow-hdr);
  border-bottom: 3px solid var(--gold);
  position: relative;
  overflow: hidden;
}}
header::before {{
  content: '';
  position: absolute; inset: 0;
  background: repeating-linear-gradient(-50deg, transparent, transparent 20px,
    rgba(255,255,255,0.012) 20px, rgba(255,255,255,0.012) 21px);
  pointer-events: none;
}}
.header-inner {{
  display: flex;
  align-items: center;
  gap: 20px;
  padding: 18px 32px;
  max-width: 1280px;
  margin: 0 auto;
  position: relative;
}}
.header-crest {{
  flex-shrink: 0;
  filter: drop-shadow(0 3px 8px rgba(0,0,0,0.55));
}}
.header-text {{ flex: 1; }}
.header-command {{
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 3px;
  color: var(--gold-light);
  text-transform: uppercase;
  margin-bottom: 2px;
}}
.header-title {{
  font-family: var(--font-ui);
  font-size: 26px;
  font-weight: 700;
  color: var(--white);
  line-height: 1.15;
}}
.header-subtitle {{
  font-size: 13px;
  color: rgba(255,255,255,0.82);
  margin-top: 5px;
  letter-spacing: 0.4px;
}}
.header-meta {{
  text-align: right;
  font-family: var(--font-mono);
  font-size: 11px;
  color: var(--gold-light);
  line-height: 1.85;
}}
.header-meta .badge {{
  display: inline-block;
  background: var(--gold);
  color: var(--navy-dark);
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1.5px;
  padding: 3px 10px;
  border-radius: var(--radius);
  margin-bottom: 5px;
  text-transform: uppercase;
}}
.header-strip {{
  background: var(--navy-dark);
  border-bottom: 1px solid rgba(200,151,26,0.25);
  padding: 4px 32px;
}}
.header-strip-inner {{
  max-width: 1280px;
  margin: 0 auto;
  display: flex;
  justify-content: space-between;
  font-family: var(--font-ui);
  font-size: 10px;
  letter-spacing: 1.3px;
  color: rgba(200,151,26,0.65);
  text-transform: uppercase;
}}

/* ── Layout ──────────────────────────────────────────────────────── */
main {{
  max-width: 1280px;
  margin: 0 auto;
  padding: 28px 32px 56px;
}}

/* ── Section headers ─────────────────────────────────────────────── */
.section-hdr {{
  display: flex;
  align-items: center;
  gap: 12px;
  margin: 32px 0 16px;
  padding: 0 0 12px 12px;
  border-bottom: 2px solid var(--gold-dark);
  border-left: 5px solid var(--navy);
}}
.section-badge {{
  background: var(--navy);
  color: var(--gold);
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 2px;
  padding: 4px 12px;
  border-radius: var(--radius);
  border: 1px solid var(--gold-dark);
  white-space: nowrap;
  text-transform: uppercase;
}}
.section-title {{
  font-family: var(--font-ui);
  font-size: 18px;
  font-weight: 700;
  color: var(--navy);
}}

/* ── Stat row ────────────────────────────────────────────────────── */
.stat-row {{
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 14px;
  margin-bottom: 28px;
}}
.stat-card {{
  background: var(--white);
  border: 1px solid var(--gray-100);
  border-top: 4px solid var(--navy);
  border-radius: var(--radius);
  padding: 16px 18px;
  box-shadow: var(--shadow);
  text-align: center;
}}
.stat-val {{
  font-family: var(--font-ui);
  font-size: 32px;
  font-weight: 700;
  color: var(--navy);
  line-height: 1;
  margin-bottom: 4px;
}}
.stat-lbl {{
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1.5px;
  text-transform: uppercase;
  color: var(--gray-400);
}}

/* ── Legend ──────────────────────────────────────────────────────── */
.legend-wrap {{
  display: flex;
  flex-wrap: wrap;
  gap: 8px 16px;
  background: var(--white);
  border: 1px solid var(--gray-100);
  border-radius: var(--radius);
  padding: 12px 16px;
  box-shadow: var(--shadow);
  margin-bottom: 24px;
}}
.legend-item {{
  display: flex;
  align-items: center;
  gap: 6px;
  font-size: 12px;
  font-family: var(--font-ui);
  font-weight: 600;
  color: var(--gray-700);
}}
.legend-swatch {{
  width: 16px; height: 16px;
  border-radius: 3px;
  border: 1px solid rgba(0,0,0,0.1);
  flex-shrink: 0;
}}
.legend-sep {{
  width: 1px;
  height: 20px;
  background: var(--gray-100);
  margin: 0 4px;
}}

/* ── Calendar table ──────────────────────────────────────────────── */
.cal-wrap {{
  overflow-x: auto;
  border-radius: var(--radius);
  border: 1px solid var(--gray-100);
  box-shadow: var(--shadow);
  margin-bottom: 24px;
}}
.cal-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 13px;
}}
.cal-table thead {{
  position: sticky;
  top: 0;
  z-index: 2;
}}
.cal-table thead tr {{ background: var(--navy); }}
.cal-table thead th {{
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1.5px;
  text-transform: uppercase;
  color: var(--white);
  padding: 10px 14px;
  text-align: left;
  white-space: nowrap;
  border-right: 1px solid rgba(255,255,255,0.08);
}}
.cal-table thead th:last-child {{ border-right: none; }}
.cal-table tbody tr {{ background: var(--white); }}
.cal-table tbody tr:nth-child(even) {{ background: var(--gray-50); }}
.cal-table tbody tr.wknd {{ background: var(--weekend-bg); }}
.cal-table tbody tr.holiday {{ background: var(--holiday-bg); }}
.cal-table tbody tr:hover {{ background: var(--navy-pale); }}
.cal-table tbody td {{
  padding: 8px 14px;
  border-bottom: 1px solid var(--gray-50);
  border-right: 1px solid var(--gray-50);
  vertical-align: middle;
  white-space: nowrap;
}}
.cal-table tbody td:last-child {{ border-right: none; }}
.date-col {{ font-family: var(--font-mono); font-size: 12px; font-weight: 700; color: var(--gray-700); }}
.dow-col  {{ font-size: 11px; color: var(--gray-400); font-weight: 700; text-transform: uppercase; letter-spacing: 1px; }}
.type-badge {{
  display: inline-block;
  font-family: var(--font-ui);
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 1px;
  text-transform: uppercase;
  padding: 2px 7px;
  border-radius: 2px;
}}
.type-weekday  {{ background: var(--gray-100); color: var(--gray-600); }}
.type-weekend  {{ background: #BFDBFE; color: #1E40AF; }}
.type-holiday  {{ background: #FED7AA; color: #9A3412; }}
.dir-chip {{
  display: inline-block;
  color: #fff;
  font-family: var(--font-ui);
  font-size: 11px;
  font-weight: 700;
  letter-spacing: 0.5px;
  padding: 4px 12px;
  border-radius: 3px;
  min-width: 72px;
  text-align: center;
}}
.month-divider td {{
  background: var(--navy-dark) !important;
  color: var(--gold-light) !important;
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 3px;
  text-transform: uppercase;
  padding: 5px 14px !important;
  border-bottom: 1px solid rgba(200,151,26,0.3) !important;
}}

/* ── Summary & fairness tables ───────────────────────────────────── */
.table-card {{
  background: var(--white);
  border: 1px solid var(--gray-100);
  border-radius: var(--radius);
  box-shadow: var(--shadow);
  overflow: hidden;
  margin-bottom: 20px;
}}
.table-card-hdr {{
  background: var(--navy);
  color: var(--white);
  font-family: var(--font-ui);
  font-size: 11px;
  font-weight: 700;
  letter-spacing: 1.5px;
  text-transform: uppercase;
  padding: 9px 16px;
  border-left: 4px solid var(--gold);
  display: flex;
  align-items: center;
  gap: 10px;
}}
.data-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 13px;
}}
.data-table thead tr {{ background: var(--navy-pale); }}
.data-table thead th {{
  font-family: var(--font-ui);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: 1.2px;
  text-transform: uppercase;
  color: var(--navy);
  padding: 9px 14px;
  text-align: center;
  border-bottom: 2px solid var(--gray-200);
  white-space: nowrap;
}}
.data-table thead th:first-child {{ text-align: left; }}
.data-table tbody tr {{ background: var(--white); }}
.data-table tbody tr:nth-child(even) {{ background: var(--gray-50); }}
.data-table tbody tr:hover {{ background: var(--navy-pale); }}
.data-table tbody td {{
  padding: 8px 14px;
  border-bottom: 1px solid var(--gray-50);
  text-align: center;
  vertical-align: middle;
}}
.data-table tbody td:first-child {{ text-align: left; font-weight: 700; }}
.bar-wrap {{ display: flex; align-items: center; gap: 8px; }}
.bar-bg {{ background: var(--gray-100); border-radius: 3px; height: 14px; flex: 1; min-width: 80px; overflow: hidden; }}
.bar-fill {{ height: 100%; border-radius: 3px; }}
.bar-num {{ font-family: var(--font-mono); font-size: 12px; font-weight: 700; min-width: 26px; text-align: right; color: var(--gray-700); }}

/* ── Gini / fairness ─────────────────────────────────────────────── */
.gini-strip {{
  display: flex;
  gap: 24px;
  padding: 14px 16px;
  border-bottom: 1px solid var(--gray-100);
  background: var(--navy-pale);
  flex-wrap: wrap;
}}
.gini-block {{ font-size: 12px; }}
.gini-val {{ font-family: var(--font-ui); font-size: 22px; font-weight: 700; color: var(--navy); }}
.gini-lbl {{ font-family: var(--font-ui); font-size: 10px; font-weight: 700; letter-spacing: 1px; text-transform: uppercase; color: var(--gray-400); }}
.pill {{
  display: inline-block;
  padding: 2px 8px;
  border-radius: 99px;
  font-family: var(--font-ui);
  font-size: 11px;
  font-weight: 700;
}}
.pill-green  {{ background: var(--green-bg);  color: var(--green-ok); }}
.pill-yellow {{ background: var(--yellow-bg); color: var(--yellow-ok); }}
.pill-red    {{ background: var(--red-bg);    color: var(--red-ok); }}
.gini-guide {{
  font-family: var(--font-mono);
  font-size: 10px;
  color: var(--gray-400);
  align-self: center;
  line-height: 1.7;
}}

/* ── Roles grid (side-by-side) ───────────────────────────────────── */
.roles-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(420px, 1fr));
  gap: 20px;
}}

/* ── Footer ──────────────────────────────────────────────────────── */
footer {{
  background: var(--navy-dark);
  color: var(--gold-light);
  font-family: var(--font-mono);
  font-size: 11px;
  text-align: center;
  padding: 14px 32px;
  border-top: 2px solid var(--gold-dark);
  line-height: 2;
}}
footer strong {{ color: var(--white); }}

/* ── Print ───────────────────────────────────────────────────────── */
@media print {{
  .stat-row {{ grid-template-columns: repeat(4, 1fr); }}
  header, .cls-banner {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
  .cal-table thead {{ position: static; }}
}}
@media (max-width: 768px) {{
  .stat-row {{ grid-template-columns: repeat(2, 1fr); }}
  main {{ padding: 16px; }}
  .header-inner {{ padding: 14px 16px; gap: 12px; }}
  .header-title {{ font-size: 20px; }}
  .header-meta {{ display: none; }}
}}
</style>
</head>
<body>

<div class="cls-banner">UNCLASSIFIED</div>

<header>
  <div class="header-inner">
    <div class="header-crest">{_USAREUR_SVG}</div>
    <div class="header-text">
      <div class="header-command">Headquarters &bull; United States Army Europe and Africa</div>
      <div class="header-title">HHBn Staff Duty Roster</div>
      <div class="header-subtitle">USAREUR-AF Operational Data Team &nbsp;&mdash;&nbsp; Quarterly Duty Assignment</div>
    </div>
    <div class="header-meta">
      <div class="badge">STAFF DUTY</div><br/>
      QUARTER: {quarter_label}<br/>
      GENERATED: {generated}<br/>
      DIST: UNLIMITED
    </div>
  </div>
</header>

<div class="header-strip">
  <div class="header-strip-inner">
    <span>HHBn &bull; SDNCO &bull; SD RUNNER &bull; QUARTERLY ASSIGNMENT</span>
    <span>ILP SOLVER &bull; PROPORTIONAL FAIRNESS &bull; UNCLASSIFIED</span>
  </div>
</div>

<main>

  <div class="section-hdr">
    <span class="section-badge">OVERVIEW</span>
    <span class="section-title">Quarter Summary &mdash; {quarter_label}</span>
  </div>

  {stat_cards}

  <div class="section-hdr">
    <span class="section-badge">LEGEND</span>
    <span class="section-title">Directorate Color Key</span>
  </div>
  {legend_html}

  <div class="section-hdr">
    <span class="section-badge">SUMMARY</span>
    <span class="section-title">Directorate Summary by Role</span>
  </div>
  {summary_html}

  <div class="section-hdr">
    <span class="section-badge">FAIRNESS</span>
    <span class="section-title">Equity Analysis &mdash; Target vs. Actual</span>
  </div>
  {fairness_html}

  <div class="section-hdr">
    <span class="section-badge">ROSTER</span>
    <span class="section-title">Daily Duty Assignment</span>
  </div>
  {calendar_html}

</main>

<footer>
  <strong>HEADQUARTERS, UNITED STATES ARMY EUROPE AND AFRICA</strong><br/>
  HHBn STAFF DUTY ROSTER &nbsp;&bull;&nbsp; SDNCO &amp; SD RUNNER &nbsp;&bull;&nbsp; {quarter_label}<br/>
  ILP PROPORTIONAL FAIRNESS SOLVER &nbsp;&bull;&nbsp; USAREUR-AF OPERATIONAL DATA TEAM<br/>
  GENERATED: {generated} &nbsp;&bull;&nbsp; UNCLASSIFIED
</footer>

<div class="cls-banner">UNCLASSIFIED</div>

</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")


# ── HTML sub-builders ─────────────────────────────────────────────────────────

def _build_legend_html(all_dirs: List[str], color_map: Dict[str, str]) -> str:
    items = "".join(
        f'<div class="legend-item">'
        f'<div class="legend-swatch" style="background:#{color_map[dn]}"></div>'
        f'<span>{dn}</span></div>'
        for dn in all_dirs
    )
    sep = '<div class="legend-sep"></div>'
    items += (
        sep +
        '<div class="legend-item">'
        '<div class="legend-swatch" style="background:#DBEAFE;border-color:#93C5FD"></div>'
        '<span>Weekend</span></div>'
        '<div class="legend-item">'
        '<div class="legend-swatch" style="background:#FCE4D6;border-color:#FDBA74"></div>'
        '<span>Holiday / Training Holiday</span></div>'
    )
    return f'<div class="legend-wrap">{items}</div>'


def _build_calendar_table_html(
    solutions: List[RosterSolution],
    all_days: List[date],
    holiday_dates: Set[date],
    color_map: Dict[str, str],
) -> str:
    """Single tabular calendar: one row per day, one column per role."""
    dow_abbr = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

    # Column headers
    role_headers = "".join(
        f'<th>{sol.role.replace("_", " ")}</th>'
        for sol in solutions
    )

    rows = ""
    prev_month = None
    for day in all_days:
        day_type = classify_day(day, holiday_dates)

        # Month divider row
        month_label = day.strftime("%B %Y").upper()
        if day.month != prev_month:
            rows += (
                f'<tr class="month-divider">'
                f'<td colspan="{3 + len(solutions)}">{month_label}</td>'
                f'</tr>'
            )
            prev_month = day.month

        # Row CSS class
        if day_type == HOLIDAY:
            row_cls = "holiday"
        elif day_type == WEEKEND:
            row_cls = "wknd"
        else:
            row_cls = ""

        # Day type badge
        type_badges = {
            WEEKDAY: '<span class="type-badge type-weekday">Weekday</span>',
            WEEKEND: '<span class="type-badge type-weekend">Weekend</span>',
            HOLIDAY: '<span class="type-badge type-holiday">Holiday</span>',
        }

        # Directorate chip cells
        dir_cells = ""
        for sol in solutions:
            assigned = sol.assignment.get(day, "")
            color    = color_map.get(assigned, "9CA3AF")
            if assigned:
                dir_cells += (
                    f'<td><span class="dir-chip" style="background:#{color}">{assigned}</span></td>'
                )
            else:
                dir_cells += '<td>—</td>'

        rows += (
            f'<tr class="{row_cls}">'
            f'<td class="date-col">{day.strftime("%d %b %Y")}</td>'
            f'<td class="dow-col">{dow_abbr[day.weekday()]}</td>'
            f'<td>{type_badges[day_type]}</td>'
            f'{dir_cells}'
            f'</tr>'
        )

    return (
        f'<div class="cal-wrap">'
        f'<table class="cal-table">'
        f'<thead><tr><th>Date</th><th>Day</th><th>Type</th>{role_headers}</tr></thead>'
        f'<tbody>{rows}</tbody>'
        f'</table>'
        f'</div>'
    )


def _build_summary_html(
    solutions: List[RosterSolution],
    all_days: List[date],
    color_map: Dict[str, str],
) -> str:
    n = len(all_days)
    cards = ""
    for sol in solutions:
        H = sum(s.eligible for s in sol.stats)
        max_total = max((s.total_days for s in sol.stats), default=1)
        rows = ""
        for s in sol.stats:
            color  = color_map.get(s.name, "9CA3AF")
            target = s.eligible / H * n
            pct    = s.total_days / max_total if max_total > 0 else 0
            bar    = (
                f'<div class="bar-wrap">'
                f'<div class="bar-bg"><div class="bar-fill" style="width:{pct*100:.0f}%;background:#{color}"></div></div>'
                f'<span class="bar-num">{s.total_days}</span>'
                f'</div>'
            )
            rows += (
                f'<tr>'
                f'<td style="color:#{color}">{s.name}</td>'
                f'<td>{s.eligible}</td>'
                f'<td>{target:.1f}</td>'
                f'<td style="min-width:160px">{bar}</td>'
                f'<td>{s.weekday_days}</td>'
                f'<td>{s.weekend_days}</td>'
                f'<td>{s.holiday_days}</td>'
                f'<td>{s.hard_days}</td>'
                f'</tr>'
            )
        cards += (
            f'<div class="table-card">'
            f'<div class="table-card-hdr">{sol.role.replace("_", " ")} &nbsp;&mdash;&nbsp; {sol.solver_status}</div>'
            f'<table class="data-table"><thead><tr>'
            f'<th>Directorate</th><th>Eligible</th><th>Target</th><th>Actual Days</th>'
            f'<th>Weekday</th><th>Weekend</th><th>Holiday</th><th>Hard Days</th>'
            f'</tr></thead><tbody>{rows}</tbody></table>'
            f'</div>'
        )
    return f'<div class="roles-grid">{cards}</div>'


def _build_fairness_html(solutions: List[RosterSolution]) -> str:
    """
    Executive-friendly fairness section.

    Leads with a plain-language verdict, shows a simple color-coded table of
    who is over/under their fair share, and explains the math in one sentence.
    No Gini jargon on the face — it's tucked into a 'How we measure this' note.
    """
    cards = ""
    for sol in solutions:
        H      = sum(s.eligible for s in sol.stats)
        n      = len(sol.assignment)
        n_hard = sum(s.hard_days for s in sol.stats)

        # ── Compute deltas ────────────────────────────────────────────────────
        def delta_class(actual: int, target: float) -> str:
            d = abs(actual - target)
            return "pill-green" if d <= 1 else ("pill-yellow" if d <= 2 else "pill-red")

        def delta_text(actual: int, target: float) -> str:
            d = actual - target
            sign = "+" if d >= 0 else ""
            if abs(d) <= 1:
                return "On target"
            elif d > 0:
                return f"+{d:.0f} over"
            else:
                return f"{d:.0f} under"

        # Overall verdict: are all directorates within ±1 of their targets?
        all_ok = all(
            abs(s.total_days - s.eligible / H * n) <= 1 and
            abs(s.hard_days  - s.eligible / H * n_hard) <= 1
            for s in sol.stats
        )
        any_red = any(
            abs(s.total_days - s.eligible / H * n) > 2 or
            abs(s.hard_days  - s.eligible / H * n_hard) > 2
            for s in sol.stats
        )

        if all_ok:
            verdict_cls  = "verdict-green"
            verdict_icon = "&#10003;"
            verdict_text = "All directorates are within one day of their fair share."
        elif any_red:
            verdict_cls  = "verdict-red"
            verdict_icon = "&#9888;"
            verdict_text = "One or more directorates are more than 2 days off their fair share. Review highlighted rows."
        else:
            verdict_cls  = "verdict-yellow"
            verdict_icon = "&#9679;"
            verdict_text = "Minor variation detected (&le;2 days). Distribution is acceptable but not perfectly equal."

        rows = ""
        for s in sol.stats:
            t_total = s.eligible / H * n
            t_hard  = s.eligible / H * n_hard if n_hard > 0 else 0
            total_cls = delta_class(s.total_days, t_total)
            hard_cls  = delta_class(s.hard_days, t_hard)
            rows += (
                f'<tr>'
                f'<td style="font-weight:700">{s.name}</td>'
                f'<td>{s.eligible} soldiers</td>'
                f'<td style="font-family:var(--font-mono)">{t_total:.1f}</td>'
                f'<td><span class="pill {total_cls}">{s.total_days} days &nbsp; {delta_text(s.total_days, t_total)}</span></td>'
                f'<td style="font-family:var(--font-mono)">{t_hard:.1f}</td>'
                f'<td><span class="pill {hard_cls}">{s.hard_days} days &nbsp; {delta_text(s.hard_days, t_hard)}</span></td>'
                f'</tr>'
            )

        # Hard-day explanation for layman
        hard_pct = round(n_hard / n * 100) if n else 0

        cards += f"""
<div class="table-card">
  <div class="table-card-hdr">{sol.role.replace("_", " ")} &nbsp;&mdash;&nbsp; Equity Report</div>

  <div class="verdict-banner {verdict_cls}">
    <span class="verdict-icon">{verdict_icon}</span>
    <span>{verdict_text}</span>
  </div>

  <div class="fairness-explainer">
    <strong>How fair shares are calculated:</strong>
    Each directorate&rsquo;s target is proportional to how many soldiers it has available for duty.
    A directorate with twice as many eligible soldiers pulls roughly twice as many duty days &mdash; including its
    proportional share of weekends and holidays ({hard_pct}% of the quarter).
    Green = within 1 day of fair share. Yellow = 1&ndash;2 days off. Red = more than 2 days off.
  </div>

  <table class="data-table">
    <thead>
      <tr>
        <th>Directorate</th>
        <th>Eligible Soldiers</th>
        <th>Fair Share (Total Days)</th>
        <th>Actual Total Days</th>
        <th>Fair Share (Wknd / Holiday)</th>
        <th>Actual Wknd / Holiday Days</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>

  <div class="gini-footnote">
    Distribution score (Gini coefficient): Total days = {sol.total_day_gini:.3f} &nbsp;&bull;&nbsp;
    Hard days = {sol.hard_day_gini:.3f} &nbsp;&bull;&nbsp;
    0.000 = perfectly equal &nbsp; &#8250; &nbsp; &le;0.050 = acceptable &nbsp; &#8250; &nbsp; &gt;0.100 = review needed
  </div>
</div>"""

    # Inject the additional CSS needed for fairness section
    css_inject = """<style>
.verdict-banner {
  display: flex; align-items: center; gap: 12px;
  padding: 12px 16px; font-family: var(--font-ui); font-size: 13px; font-weight: 600;
  border-bottom: 1px solid var(--gray-100);
}
.verdict-icon { font-size: 18px; flex-shrink: 0; }
.verdict-green  { background: var(--green-bg);  color: var(--green-ok);  }
.verdict-yellow { background: var(--yellow-bg); color: var(--yellow-ok); }
.verdict-red    { background: var(--red-bg);    color: var(--red-ok);    }
.fairness-explainer {
  padding: 10px 16px 12px;
  font-size: 12px;
  color: var(--gray-600);
  line-height: 1.65;
  border-bottom: 1px solid var(--gray-100);
  background: var(--off-white);
}
.gini-footnote {
  padding: 8px 16px;
  font-family: var(--font-mono);
  font-size: 10px;
  color: var(--gray-400);
  letter-spacing: 0.5px;
  border-top: 1px solid var(--gray-50);
}
</style>"""

    return css_inject + f'<div class="roles-grid">{cards}</div>'
